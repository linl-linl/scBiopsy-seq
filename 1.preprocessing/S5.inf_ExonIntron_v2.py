# !/usr/bin/env python
# -*- coding=utf-8 -*-

import os
import re
import openpyxl
import argparse

parser = argparse.ArgumentParser(description='整合各样本的ExonIntron信息（对目录下所有*_ExonIntron.txt操作）: \
                                              python inf_ExonIntron_v2.py \
                                              -o output')

parser.add_argument('-o', '--output', type=str, required=True, help='output_file_name')

args = parser.parse_args()

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'
ws['A1'] = 'File'
ws['B1'] = 'Sample'
ws['C1'] = '% of bases mapped to exon'
ws['D1'] = '% of bases mapped to intron'
ws['E1'] = '% of bases mapped to gene'
ws['F1'] = 'Exon bases'
ws['G1'] = 'Gene bases'
ws['H1'] = 'Total mapped bases'
count = 2
os.system("ls *_ExonIntron.txt > list_ExonIntron_tmp.txt")
file_list = open('list_ExonIntron_tmp.txt', 'r')
for i in file_list:
    m = re.sub('\n', '', i)
    file_ExonIntron = open(m, 'r')
    for i2 in file_ExonIntron:
        m2 = re.sub('\n', '', i2)
        n = m2.split(': ')
        if n[0] == 'Sample name':
            Sample_name = n[1]
        elif n[0] == 'Total mapped bases':
            Total_mapped_bases = n[1]
        elif n[0] == 'Exon bases':
            Exon_bases = n[1]
        elif n[0] == '% of bases mapped to exon':
            p_of_bases_mapped_to_exon = n[1]
        elif n[0] == 'Gene bases':
            Gene_bases = n[1]
        elif n[0] == '% of bases mapped to gene':
            p_of_bases_mapped_to_gene = n[1]
        elif n[0] == '% of bases mapped to intron':
            p_of_bases_mapped_to_intron = n[1]
    ws['A'+str(count)] = Sample_name
    ws['B'+str(count)] = Sample_name.split('_')[0]
    ws['C'+str(count)] = p_of_bases_mapped_to_exon
    ws['D'+str(count)] = p_of_bases_mapped_to_intron
    ws['E'+str(count)] = p_of_bases_mapped_to_gene
    ws['F'+str(count)] = Exon_bases
    ws['G'+str(count)] = Gene_bases
    ws['H'+str(count)] = Total_mapped_bases
    count = count+1
file_list.close()
os.system("rm list_ExonIntron_tmp.txt")
wb.save(args.output+'_ExonIntron_MAP.xlsx')

