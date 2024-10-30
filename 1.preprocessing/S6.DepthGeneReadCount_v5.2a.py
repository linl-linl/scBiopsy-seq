#!/usr/bin/env python
#coding=utf-8

import os
import re
import numpy as np
import openpyxl
import argparse

parser = argparse.ArgumentParser(description='取同一测序深度，计算测到的基因数 \
                                              python DepthGeneReadCount_v3.py \
                                              -s list_Aligned.out.sam \
                                              -g url_to_gtf \
                                              -d deptha \
                                              -o output')

parser.add_argument('-s', '--sam', type=str, nargs='+', help='a sam file(s) with suffix".sam" or a list of sam')
parser.add_argument('-g', '--gtf', type=str, default='/home/disk/linl/reference/GRCh38/Homo_sapiens.GRCh38.105.gtf', help='gtf for htseq-count')
parser.add_argument('-mt', '--mtgene', type=str, default='/home/disk/linl/reference/GRCh38/mt_gene.txt', help='a list of mt gene id/name')
parser.add_argument('-rp', '--rpgene', type=str, default='/home/disk/linl/dmf_live_seq/live_seq/ncbi_rp_id.txt', help='a list of rp gene id/name')
parser.add_argument('-d', '--depth', type=str, required=True, help='depth or "raw"')
parser.add_argument('-o', '--output', type=str, required=True, help='output_depth_GeneReadCount.xlsx')

args = parser.parse_args()


def DepthGene(sam, fq, d, depth_str):
    file_fq = open(fq, 'r')
    cont_fq = file_fq.readlines()
    nn = 0
    nn_max = len(cont_fq)
    pattern = r'@(.*?)\s'
    list_id = []
    while (nn<nn_max):
        if nn%4==0:
            idid = re.match(pattern, cont_fq[nn])
            list_id.append(idid.group(1).split('/')[0])
        nn = nn+1
    list_ID = np.random.choice(list_id, d, replace = False)
    dict_ID = {}
    for i in list_id:
        dict_ID[i] = 0
    for i in list_ID:
        dict_ID[i] = 1
    file_fq.close()

    align_sam = open(sam, 'r')
    prefix = re.sub('Aligned.out.sam','',sam)
    prefix = re.sub('.sorted.sam','',prefix)
    prefix = re.sub('.sam', '', prefix)
    align_sam_ID = open(prefix+'_'+depth_str+'depth.sam','w')
    for line in align_sam:
        if line.startswith('@'):
            align_sam_ID.write(line)
        else:
            ll = line.split('\t')
            if dict_ID[ll[0]]:
                align_sam_ID.write(line)
    align_sam.close()
    align_sam_ID.close()


def GeneReadCount(htseq_out):
    file = open(htseq_out,'r')
    prefix = re.sub('_counts.txt','',htseq_out)
    file_count = open(prefix+'_GeneReadCount.txt','w')
    ncount=0
    ngene=0
    mtcount=0
    rpcount=0
    for line in file:
        if ('processed' in line) or ('__' in line):
            continue
        else:
            line2 = re.sub('\n','',line)
            l = line2.split('\t')
            if (l[1] != '0'):
                ncount=ncount+int(l[1])
                ngene=ngene+1
                file_count.write(l[0]+'\t'+l[1]+'\n')
                if l[0] in mtgene:
                    mtcount=mtcount+int(l[1])
                if l[0] in rpgene:
                    rpcount=rpcount+int(l[1])
    file_count.write('nCount: '+str(ncount)+'\n')
    file_count.write('nGene: '+str(ngene)+'\n')
    file_count.write('mtCount: '+str(mtcount)+'\n')
    file_count.write('mtProp: %.2f%%\n' % (mtcount/ncount*100))
    file_count.write('rpCount: '+str(rpcount)+'\n')
    file_count.write('rpProp: %.2f%%' % (rpcount/ncount*100))
    file.close()
    file_count.close()


## input sam file
if args.sam:
    list_sam = []
    for ss in args.sam:
        if ss.endswith('.sam'):
            list_sam.append(ss)
        else:
            file = open(ss, 'r')
            for i in file:
                list_sam.append(i.rstrip())
            file.close()
    list_sam = sorted(list(set(list_sam)))
else:
    #os.system("ls *Aligned.out.sam > list_sam_htseqS10.tmp.txt")
    #os.system("ls *.sorted.sam > list_sam_htseqS10.tmp.txt")
    os.system("ls *.sam > list_sam_htseqS10.tmp.txt")
    list_sam = []
    file = open('list_sam_htseqS10.tmp.txt', 'r')
    for i in file:
        list_sam.append(i.rstrip())
    os.system("rm list_sam_htseqS10.tmp.txt")

## depth
d = str(args.depth)
if (d=='raw'):
    downsample = False
    depth_str = 'raw'
else:
    downsample = True
    if ('M' in d) or ('m' in d):
        depth_str = re.sub('m', 'M', d)
        d = re.sub('M', '', d)
        d = re.sub('m', '', d)
        d = int(float(d)*1000000)
    else:
        d = int(d)
        depth_str = str((d/1000000)) + 'M'

## reference & mt/rp
gtf = args.gtf
mt = args.mtgene
rp = args.rpgene
mtfile = open(mt, 'r')
mtgene = [i.rstrip() for i in mtfile.readlines()]
mtfile.close()
rpfile = open(rp, 'r')
rpgene = [i.rstrip() for i in rpfile.readlines()]
rpfile.close()

output = args.output
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'
row=5
for sam in list_sam:

    print('''now processing %s''' % (sam))

    prefix = re.sub('Aligned.out.sam', '', sam)
    prefix = re.sub('.sorted.sam', '', prefix)
    prefix = re.sub('.sam', '', prefix)
    url_fq = '../'+prefix+'_1_repair_1.fq'

    if downsample:
        #### execute:downsample
        DepthGene(sam, url_fq, d, depth_str)

        ## output filename
        DepthGene_out = prefix+'_'+depth_str+'depth.sam'
        htseq_out1 = prefix+'_'+depth_str+'_union_counts.txt'
        GeneReadCount_out1 = prefix+'_'+depth_str+'_union_GeneReadCount.txt'
        #htseq_out2 = prefix+'_'+depth_str+'_nonempty_counts.txt'
        #GeneReadCount_out2 = prefix+'_'+depth_str+'_nonempty_GeneReadCount.txt'
        output_name = output+'_'+depth_str+'_GeneReadCounts.xlsx'
    else:
        ## output filename
        DepthGene_out = sam
        htseq_out1 = prefix+'_union_counts.txt'
        GeneReadCount_out1 = prefix+'_union_GeneReadCount.txt'
        #htseq_out2 = prefix+'_nonempty_counts.txt'
        #GeneReadCount_out2 = prefix+'_nonempty_GeneReadCount.txt'
        output_name = output+'_GeneReadCounts.xlsx'

    #### execute:htseq-count
    os.system("htseq-count -f sam -r name -s no -a 10 -t exon -i gene_id -m union %s %s > %s" % (DepthGene_out, gtf, htseq_out1))
    #os.system("htseq-count -f sam -r name -s no -a 10 -t gene -i gene_id -m intersection-nonempty %s %s > %s" %(DepthGene_out, gtf, htseq_out2))

    result1=os.popen("wc -l %s" %(htseq_out1)).read()
    num1=result1.split(' ')[0]
    total_gene_num1=int(num1)-5
    #result2=os.popen("wc -l %s" %(htseq_out2)).read()
    #num2=result2.split(' ')[0]
    #total_gene_num2=int(num2)-5

    # execute:calculate nCount & nGene
    GeneReadCount(htseq_out1)
    #GeneReadCount(htseq_out2)

    file_GeneReadCount1=open(GeneReadCount_out1,'r')
    cont1=file_GeneReadCount1.readlines()
    file_GeneReadCount1.close()
    ncount1 =cont1[-6].rstrip().split(': ')[1]
    ngene1  =cont1[-5].rstrip().split(': ')[1]
    mtcount1=cont1[-4].rstrip().split(': ')[1]
    mtprop1 =cont1[-3].rstrip().split(': ')[1]
    rpcount1=cont1[-2].rstrip().split(': ')[1]
    rpprop1 =cont1[-1].rstrip().split(': ')[1]
    #result2=os.popen("wc -l %s" %(htseq_out2)).read()
    #result2=os.popen("wc -l %s" %(htseq_out2)).read()
    #file_GeneReadCount2.close()
    #ncount2 =cont2[-4].rstrip().split(': ')[1]
    #ngene2  =cont2[-3].rstrip().split(': ')[1]
    #mtcount2=cont2[-2].rstrip().split(': ')[1]
    #mtprop2 =cont2[-1].rstrip().split(': ')[1]

    ws['A'+str(row)] = prefix
    ws['B'+str(row)] = prefix.split('_')[0]
    ws['C'+str(row)] = ncount1
    ws['D'+str(row)] = ngene1
    ws['E'+str(row)] = mtcount1
    ws['F'+str(row)] = mtprop1
    ws['G'+str(row)] = rpcount1
    ws['H'+str(row)] = rpprop1
    #ws['G'+str(row)] = ncount2
    #ws['H'+str(row)] = ngene2
    #ws['I'+str(row)] = mtcount2
    #ws['J'+str(row)] = mtprop2

    print(prefix+' is finished!')
    row=row+1
    
ws['A1'] = 'Depth reads'
ws['B1'] = depth_str
ws['A2'] = 'Total gene num'
ws['B2'] = str(total_gene_num1)
#if (total_gene_num1==total_gene_num2):
#    ws['A2'] = 'Total gene num'
#    ws['B2'] = str(total_gene_num1)
#else:
#    ws['A2'] = 'Total gene num(union)'
#    ws['B2'] = str(total_gene_num1)
#    ws['C2'] = 'Total gene num(nonempty)'
#    ws['D2'] = str(total_gene_num2)
ws['C3']='基因总表达量'
ws['D3']='基因数'
ws['E3']='线粒体表达量'
ws['F3']='线粒体表达量占比'
ws['G3']='核糖蛋白表达量'
ws['H3']='核糖蛋白表达量占比'
ws['A4'] = 'File'
ws['B4'] = 'Sample'
ws['C4'] = 'nCount(union)'
ws['D4'] = 'nGene(union)'
ws['E4'] = 'mtCount(union)'
ws['F4'] = 'mtProp(union)'
ws['G4'] = 'rpCount(union)'
ws['H4'] = 'rpProp(union)'
#ws['G4'] = 'nCount(nonempty)'
#ws['H4'] = 'nGene(nonempty)'
#ws['I4'] = 'mtCount(nonempty)'
#ws['J4'] = 'mtProp(nonempty)'
wb.save(output_name)

